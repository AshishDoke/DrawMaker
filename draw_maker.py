"""
TT Cybage Internal Draw Maker - 2026
Reads Self Nomination.xlsx and generates draws in the same format
as TT Cybage Internal 2025 Draws.xlsx
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime

# ──────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────
INPUT_FILE  = "Self Nomination.xlsx"
OUTPUT_FILE = "TT Cybage Internal 2026 Draws.xlsx"

# ── Draw-size helpers (power-of-2 only) ──────────────────────────
POWERS_OF_2 = [4, 8, 16, 32, 64, 128, 256]


def next_power_of_2(n):
    """Smallest power of 2 >= n (minimum 4)."""
    p = 4
    while p < n:
        p *= 2
    return p


def prev_power_of_2(n):
    """Largest power of 2 strictly less than next_power_of_2(n), minimum 4."""
    return max(4, next_power_of_2(n) // 2)


def valid_group_sizes(draw_size):
    """
    Return all valid group sizes for a given draw_size.
    Must be even, a power of 2, and divide draw_size evenly.
    """
    sizes = []
    p = 4
    while p <= draw_size:
        sizes.append(p)
        p *= 2
    return sizes

# ──────────────────────────────────────────────────────────────────
# STYLES
# ──────────────────────────────────────────────────────────────────
TITLE_FONT      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
TITLE_FILL      = PatternFill("solid", fgColor="1F4E79")     # dark blue
GROUP_FONT      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
GROUP_FILL      = PatternFill("solid", fgColor="2E75B6")     # medium blue
HEADER_FONT     = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL     = PatternFill("solid", fgColor="ED7D31")     # orange
PLAYER_FONT     = Font(name="Calibri", size=10)
PLAYER_FILL_A   = PatternFill("solid", fgColor="DEEAF1")     # light blue
PLAYER_FILL_B   = PatternFill("solid", fgColor="FFFFFF")     # white
BYE_FONT        = Font(name="Calibri", size=10, italic=True, color="808080")
SEEDED_FONT     = Font(name="Calibri", size=10, bold=True, color="1F4E79")  # dark blue bold
SEEDED_FILL_A   = PatternFill("solid", fgColor="D6E4F0")   # slightly deeper blue for seeded rows
SEEDED_FILL_B   = PatternFill("solid", fgColor="EBF3FB")
CENTER_ALIGN    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN      = Alignment(horizontal="left",  vertical="center", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


# ──────────────────────────────────────────────────────────────────
# STEP 1 – READ NOMINATIONS
# ──────────────────────────────────────────────────────────────────
def read_nominations(filepath):
    """Return dict of category → list of formatted name strings."""
    global _name_lookup
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    ms_players  = []
    md_raw      = []   # (name_A, emp_A, name_B_hint, emp_B_hint)
    mixed_raw   = []
    ws_players  = []
    wd_raw      = []

    # Build emp-code → name lookup from registered players
    emp_lookup = {}
    for row in range(2, ws.max_row + 1):
        name     = ws.cell(row, 5).value
        emp_code = str(ws.cell(row, 6).value).strip() if ws.cell(row, 6).value else ""
        if name and emp_code:
            emp_lookup[emp_code] = name

    # Build name → emp lookup for partner resolution
    _name_lookup = build_name_lookup(emp_lookup)

    for row in range(2, ws.max_row + 1):
        name        = ws.cell(row, 5).value
        emp_code    = str(ws.cell(row, 6).value).strip() if ws.cell(row, 6).value else ""
        gender      = ws.cell(row, 7).value
        ms_yes      = ws.cell(row, 8).value   == "Yes"
        md_yes      = ws.cell(row, 9).value   == "Yes"
        mixed_yes   = ws.cell(row, 10).value  == "Yes"
        ws_yes      = ws.cell(row, 11).value  == "Yes"
        wd_yes      = ws.cell(row, 12).value  == "Yes"
        md_partner  = str(ws.cell(row, 13).value).strip() if ws.cell(row, 13).value else ""
        wd_partner  = str(ws.cell(row, 14).value).strip() if ws.cell(row, 14).value else ""
        mx_partner  = str(ws.cell(row, 15).value).strip() if ws.cell(row, 15).value else ""

        if not name:
            continue

        label = f"{name} ({emp_code})"

        if ms_yes:
            ms_players.append({"name": name, "emp": emp_code, "label": label})

        if md_yes:
            md_raw.append({
                "name": name, "emp": emp_code,
                "partner_hint": md_partner
            })

        if mixed_yes:
            mixed_raw.append({
                "name": name, "emp": emp_code,
                "partner_hint": mx_partner
            })

        if ws_yes:
            ws_players.append({"name": name, "emp": emp_code, "label": label})

        if wd_yes:
            wd_raw.append({
                "name": name, "emp": emp_code,
                "partner_hint": wd_partner
            })

    return {
        "MS":    ms_players,
        "MD":    deduplicate_pairs(md_raw,    "MD",    emp_lookup),
        "Mixed": deduplicate_pairs(mixed_raw, "Mixed", emp_lookup),
        "WS":    ws_players,
        "WD":    deduplicate_pairs(wd_raw,    "WD",    emp_lookup),
    }, emp_lookup


# ──────────────────────────────────────────────────────────────────
# STEP 2 – DEDUPLICATE DOUBLES PAIRS
# ──────────────────────────────────────────────────────────────────
def build_name_lookup(emp_lookup):
    """
    Build multiple name → emp_code mappings for fuzzy name matching.
    emp_lookup is {emp_code: full_name}.
    Returns dict of normalised_key → emp_code.
    """
    lk = {}
    for emp, full_name in emp_lookup.items():
        words = full_name.lower().split()
        # Full name
        lk[full_name.lower()] = emp
        # First + last
        if len(words) >= 2:
            lk[f"{words[0]} {words[-1]}"] = emp
        # Last name only (if unique enough – length > 4)
        if words and len(words[-1]) > 4:
            lk.setdefault(words[-1], emp)
        # First name + second word
        if len(words) >= 3:
            lk.setdefault(f"{words[0]} {words[1]}", emp)
    return lk


# Global; populated inside read_nominations before dedup
_name_lookup: dict = {}


def resolve_partner_name(hint, emp_lookup):
    """
    Try to resolve a partner hint to (resolved_name, emp_code).
    Handles: pure emp code, embedded (XXXXX) in name, or free-text name.
    """
    if not hint:
        return None, None
    hint = hint.strip()

    # 1. Pure number
    clean = hint.replace("-", "").replace(" ", "")
    if clean.isdigit():
        name = emp_lookup.get(clean)
        if name:
            return name, clean
        # Emp code provided but not in our registrations
        return f"(Emp.{clean} – unregistered)", clean

    # 2. Scan tokens for 4-6 digit emp codes (handles "Name (12345)" etc.)
    import re
    # Also strip email addresses – replace with empty
    hint_clean = re.sub(r'\S+@\S+', '', hint).strip()
    found_codes = re.findall(r'\b(\d{4,6})\b', hint_clean)
    for code in found_codes:
        if code in emp_lookup:
            return emp_lookup[code], code
        # Found a numeric code but not registered
        if code:
            return f"(Emp.{code} – unregistered)", code

    # 3. Try name-based lookup using _name_lookup (fuzzy)
    norm = hint.lower().strip()
    if norm in _name_lookup:
        emp = _name_lookup[norm]
        return emp_lookup[emp], emp
    # Try just significant words (ignore middle names)
    words = norm.split()
    if len(words) >= 2:
        short = f"{words[0]} {words[-1]}"
        if short in _name_lookup:
            emp = _name_lookup[short]
            return emp_lookup[emp], emp
        # Try first word only as last resort
        first = words[0]
        if first in _name_lookup:
            emp = _name_lookup[first]
            return emp_lookup[emp], emp

    # 4. Unresolved – return hint as-is
    return hint, None


def make_pair_key(emp_a, emp_b):
    """Canonical pair key regardless of order."""
    a = str(emp_a).strip()
    b = str(emp_b).strip() if emp_b else ""
    return tuple(sorted([a, b]))


def deduplicate_pairs(raw_list, category, emp_lookup):
    """
    Identify unique pairs from raw registration data.
    Both partners often register separately – collapse them into one entry.
    Uses resolved emp codes as keys wherever possible so cross-registration
    (A writes emp code, B writes name) still deduplicates correctly.
    Returns list of pair dicts with 'label', 'player_a', 'player_b'.
    """
    seen_keys    = set()
    unique_pairs = []
    unresolved   = []

    for entry in raw_list:
        emp_a  = entry["emp"]
        name_a = entry["name"]
        partner_hint = entry.get("partner_hint", "")

        partner_name, emp_b = resolve_partner_name(partner_hint, emp_lookup)

        # Build the canonical pair key
        if emp_b and emp_b in emp_lookup:
            # Both emps known → best key
            key = make_pair_key(emp_a, emp_b)
        elif emp_b:
            # emp_b known but not in our registrations (external partner)
            key = make_pair_key(emp_a, emp_b)
        else:
            # Fallback: sorted lower-case names
            pb = str(partner_name or "").lower().strip()
            pa = name_a.lower().strip()
            key = tuple(sorted([pa, pb]))

        if key in seen_keys:
            continue  # duplicate – skip

        seen_keys.add(key)

        if partner_name and str(partner_name).upper() != "TBD":
            label_b = (f"{partner_name} ({emp_b})"
                       if emp_b and emp_b in emp_lookup
                       else str(partner_name))
            label = f"{name_a} ({emp_a})  &  {label_b}"
        else:
            label = f"{name_a} ({emp_a})  &  TBD"
            unresolved.append(label)

        unique_pairs.append({
            "label":        label,
            "player_a":     f"{name_a} ({emp_a})",
            "partner_hint": partner_hint,
        })

    if unresolved:
        print(f"\n  [{category}] Note: {len(unresolved)} pair(s) with TBD partner "
              f"(partner did not fill form or name unclear):")
        for u in unresolved:
            print(f"    - {u}")

    return unique_pairs


# ──────────────────────────────────────────────────────────────────
# STEP 3 – BUILD GROUPS
# ──────────────────────────────────────────────────────────────────
def make_draw_groups(players, draw_size, group_size):
    """
    Shuffle players, fill to draw_size with BYEs, split into even groups.

    draw_size  – total bracket slots (must be a power of 2).
    group_size – players per round-robin group (even, power of 2,
                 and must divide draw_size exactly).

    Returns:
        groups     – list of lists, each exactly group_size long.
        waitlisted – players beyond draw_size (coordinator decides fate).
    """
    if draw_size % group_size != 0:
        raise ValueError(
            f"group_size ({group_size}) must divide draw_size ({draw_size}) exactly."
        )

    shuffled = list(players)
    random.shuffle(shuffled)

    n = len(shuffled)

    # Players beyond the draw size go to the waitlist
    if n > draw_size:
        waitlisted = shuffled[draw_size:]
        active     = shuffled[:draw_size]
    else:
        waitlisted = []
        active     = shuffled[:]

    # Pad active list with BYEs to fill all draw slots
    byes_needed = draw_size - len(active)
    bye_entry   = {"label": "--BYE--", "is_bye": True}
    all_slots   = active + [bye_entry] * byes_needed

    # Split evenly into groups of group_size
    num_groups = draw_size // group_size
    groups = [
        all_slots[i * group_size: (i + 1) * group_size]
        for i in range(num_groups)
    ]

    return groups, waitlisted


# ──────────────────────────────────────────────────────────────────
# STEP 4 – WRITE EXCEL SHEET
# ──────────────────────────────────────────────────────────────────
def write_draw_sheet(ws, title, groups, waitlisted=None):
    """
    Write a category draw sheet matching TT 2025 format.
    Columns:
      A  = player number
      B  = player / pair name (wide)
      C-G = spacer
      H  = QF result placeholder
      I  = spacer (header: QF)
      J  = SF result placeholder
      K  = spacer (header: SF)
      L  = Finals placeholder   (header: Finals in col M)
    """
    # ── column widths ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 5
    ws.column_dimensions["J"].width = 33
    ws.column_dimensions["K"].width = 5
    ws.column_dimensions["L"].width = 5
    ws.column_dimensions["M"].width = 36

    # ── row 1 – title + stage headers ─────────────────────────────
    ws.row_dimensions[1].height = 24
    title_cell = ws.cell(1, 1, title)
    title_cell.font  = TITLE_FONT
    title_cell.fill  = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.merge_cells("A1:G1")

    for col, label in [(9, "QF"), (11, "SF"), (13, "Finals")]:
        c = ws.cell(1, col, label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN

    current_row   = 3
    player_number = 1

    # Track bracket row placements for each group
    group_mid_rows = []   # SF level – middle row of each group span
    qf_rows_all   = []    # [(qf1_row, qf2_row), ...] per group

    for g_idx, group in enumerate(groups):
        group_start  = current_row
        n_in_group   = len(group)

        # ── group header ──────────────────────────────────────────
        header_text = f"------ GROUP - {g_idx + 1} -----"
        ws.row_dimensions[current_row].height = 18
        hdr = ws.cell(current_row, 1, header_text)
        hdr.font      = GROUP_FONT
        hdr.fill      = GROUP_FILL
        hdr.alignment = LEFT_ALIGN
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=7
        )
        current_row += 1
        blank_row_after_header = current_row
        current_row += 1  # blank row

        # ── player rows ───────────────────────────────────────────
        first_player_row = current_row
        for p_idx, player in enumerate(group):
            ws.row_dimensions[current_row].height = 17

            label = player.get("label", player) if isinstance(player, dict) else str(player)
            is_bye    = player.get("is_bye",    False) if isinstance(player, dict) else False
            is_seeded = player.get("is_seeded", False) if isinstance(player, dict) else False

            num_cell  = ws.cell(current_row, 1, "" if is_bye else player_number)
            name_cell = ws.cell(current_row, 2, label)

            if is_bye:
                fill = PatternFill("solid", fgColor="F2F2F2")
                font = BYE_FONT
            elif is_seeded:
                fill = SEEDED_FILL_A if p_idx % 2 == 0 else SEEDED_FILL_B
                font = SEEDED_FONT
            else:
                fill = PLAYER_FILL_A if p_idx % 2 == 0 else PLAYER_FILL_B
                font = PLAYER_FONT

            for c in [num_cell, name_cell]:
                c.fill      = fill
                c.border    = thin_border()
                c.alignment = CENTER_ALIGN if c.column == 1 else LEFT_ALIGN

            name_cell.font = font
            num_cell.font  = Font(name="Calibri", size=10, bold=True)

            if not is_bye:
                player_number += 1
            current_row += 1

        last_player_row = current_row - 1
        group_mid = (first_player_row + last_player_row) // 2

        # ── QF bracket placeholders (per half-group) ──────────────
        if n_in_group >= 4:
            half        = n_in_group // 2
            qf1_row     = (first_player_row + first_player_row + half - 1) // 2
            qf2_row     = (first_player_row + half + last_player_row) // 2
        else:
            qf1_row = group_mid
            qf2_row = group_mid

        qf_rows_all.append((qf1_row, qf2_row))

        # QF placeholder cells
        for qf_row in set([qf1_row, qf2_row]):
            c = ws.cell(qf_row, 8)
            c.border    = Border(bottom=Side(style="medium", color="2E75B6"))
            c.alignment = LEFT_ALIGN
            c.font      = Font(name="Calibri", size=10, italic=True, color="808080")

        # SF placeholder cell
        sf_cell = ws.cell(group_mid, 10)
        sf_cell.border    = Border(bottom=Side(style="medium", color="ED7D31"))
        sf_cell.alignment = LEFT_ALIGN
        sf_cell.font      = Font(name="Calibri", size=10, italic=True, color="808080")

        group_mid_rows.append(group_mid)

        # ── blank row after group ─────────────────────────────────
        current_row += 1  # blank separator row
        group_end = current_row - 1

    # ── Finals placeholder ─────────────────────────────────────────
    if group_mid_rows:
        finals_row = (group_mid_rows[0] + group_mid_rows[-1]) // 2
        f_cell = ws.cell(finals_row, 13)
        f_cell.border    = Border(bottom=Side(style="thick", color="1F4E79"))
        f_cell.alignment = LEFT_ALIGN
        f_cell.font      = Font(name="Calibri", size=10, bold=True, italic=True, color="1F4E79")

    # ── Waitlisted section (players beyond the chosen draw size) ────
    if waitlisted:
        current_row += 2  # blank separator rows

        ws.row_dimensions[current_row].height = 20
        wl_hdr = ws.cell(current_row, 1,
                         "── WAITLISTED / RESERVE PLAYERS ──")
        wl_hdr.font      = Font(name="Calibri", bold=True, size=11,
                                color="FFFFFF")
        wl_hdr.fill      = PatternFill("solid", fgColor="595959")
        wl_hdr.alignment = LEFT_ALIGN
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=7)
        current_row += 1

        ws.row_dimensions[current_row].height = 14
        note_cell = ws.cell(current_row, 2,
                            "These players exceed the draw size. "
                            "Tournament coordinator to decide eligibility.")
        note_cell.font      = Font(name="Calibri", size=9, italic=True,
                                   color="595959")
        note_cell.alignment = LEFT_ALIGN
        current_row += 1

        for wl_player in waitlisted:
            ws.row_dimensions[current_row].height = 17
            lbl = (wl_player.get("label", str(wl_player))
                   if isinstance(wl_player, dict) else str(wl_player))
            c = ws.cell(current_row, 2, lbl)
            c.font      = Font(name="Calibri", size=10, italic=True,
                               color="595959")
            c.fill      = PatternFill("solid", fgColor="FFF2CC")
            c.border    = thin_border()
            c.alignment = LEFT_ALIGN
            current_row += 1

    # ── freeze top row ─────────────────────────────────────────────
    ws.freeze_panes = "A2"

    return current_row


# ──────────────────────────────────────────────────────────────────
# STEP 5 – GENERATE SUMMARY SHEET
# ──────────────────────────────────────────────────────────────────
def write_summary_sheet(ws, category_counts):
    """
    category_counts: list of
        (title, total_entries, draw_size, num_groups, byes, waitlisted_count)
    """
    col_widths = [32, 12, 12, 10, 8, 14]
    col_headers = ["Category", "Entries", "Draw Size", "Groups",
                   "BYEs", "Waitlisted"]
    for i, (w, _) in enumerate(zip(col_widths, col_headers), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    title_cell = ws.cell(1, 1, "TT Cybage Internal 2026 – Draw Summary")
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")

    ws.row_dimensions[2].height = 18
    for col, txt in enumerate(col_headers, start=1):
        c = ws.cell(2, col, txt)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN

    for r, row_data in enumerate(category_counts, start=3):
        title_val, entries, draw_size, n_groups, byes, waitlisted_count = row_data
        ws.row_dimensions[r].height = 17
        fill = PLAYER_FILL_A if r % 2 == 0 else PLAYER_FILL_B
        vals = [title_val, entries, draw_size, n_groups, byes, waitlisted_count]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(r, col, val)
            c.font      = PLAYER_FONT
            c.fill      = fill
            c.alignment = CENTER_ALIGN
            c.border    = thin_border()

    note_row = len(category_counts) + 5
    ws.row_dimensions[note_row].height = 15
    note = ws.cell(note_row, 1,
                   f"Draw generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}")
    note.font = Font(name="Calibri", size=9, italic=True, color="808080")


# ──────────────────────────────────────────────────────────────────
# GENERATE EXCEL  (called from both the UI and CLI)
# ──────────────────────────────────────────────────────────────────
def generate_draw_excel(draw_configs):
    """
    Build the draw Excel workbook and return its content as bytes.

    draw_configs: dict  cat → {
        "entries":    list of player/pair dicts (each must have a 'label' key),
        "draw_size":  int  – power of 2 (e.g. 64, 128),
        "group_size": int  – even, power of 2, must divide draw_size,
    }
    Returns bytes of the generated Excel file.
    """
    wb = openpyxl.Workbook()
    sheet_cfg = [
        ("MS",    "TT 2026 Mens Singles"),
        ("MD",    "TT 2026 Mens Doubles"),
        ("Mixed", "TT 2026 Mixed Doubles"),
        ("WS",    "TT 2026 Womens Singles"),
        ("WD",    "TT 2026 Womens Doubles"),
    ]

    category_counts = []
    first = True

    for cat, title in sheet_cfg:
        if cat not in draw_configs or not draw_configs[cat].get("entries"):
            continue

        cfg        = draw_configs[cat]
        entries    = cfg["entries"]
        draw_size  = cfg["draw_size"]
        group_size = cfg["group_size"]

        # Normalise to simple label dicts (works for both singles and doubles)
        players = [{"label": p["label"]} for p in entries]

        groups, waitlisted = make_draw_groups(players, draw_size, group_size)

        if first:
            ws = wb.active
            ws.title = cat
            first = False
        else:
            ws = wb.create_sheet(title=cat)

        write_draw_sheet(ws, title, groups, waitlisted=waitlisted)

        byes = draw_size - min(len(players), draw_size)
        category_counts.append((
            title,
            len(entries),
            draw_size,
            len(groups),
            byes,
            len(waitlisted),
        ))

    sum_ws = wb.create_sheet(title="Summary")
    write_summary_sheet(sum_ws, category_counts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# SEEDING – READ EXISTING DRAW FILE
# ──────────────────────────────────────────────────────────────────

# Sheet names as generated by this tool
_DRAW_SHEET_CATS = ["MS", "MD", "Mixed", "WS", "WD"]


def read_draw_file(filepath_or_bytes):
    """
    Parse an existing draw Excel (generated by this app).
    Detects player names in **bold** as seeded players.
    Seeds are numbered in the order they appear top-to-bottom.
    Also detects draw_size and group_size from the sheet structure.

    Returns:
        dict  cat → {"players": [...], "seeded": [...],
                      "draw_size": int, "group_size": int}
        Each player dict: {"label": str, "is_seeded": bool}
    """
    wb = openpyxl.load_workbook(filepath_or_bytes)
    result = {}

    for cat in _DRAW_SHEET_CATS:
        if cat not in wb.sheetnames:
            result[cat] = {"players": [], "seeded": [], "draw_size": 0, "group_size": 0}
            continue

        ws = wb[cat]
        players = []
        seeded  = []
        in_waitlist    = False
        group_sizes    = []
        cur_group_slots = 0

        for row in range(2, ws.max_row + 1):
            cell_a = ws.cell(row, 1)
            cell_b = ws.cell(row, 2)
            val_a  = str(cell_a.value or "").strip()
            val_b  = str(cell_b.value or "").strip()

            # Detect waitlisted section start
            if "WAITLISTED" in val_a.upper():
                if cur_group_slots:
                    group_sizes.append(cur_group_slots)
                    cur_group_slots = 0
                in_waitlist = True
                continue

            # Detect group header (col A holds the header text; col B is merged/empty)
            if val_a.startswith("------"):
                if cur_group_slots:
                    group_sizes.append(cur_group_slots)
                    cur_group_slots = 0
                continue

            # Skip empty rows and notes
            if not val_b:
                continue
            if "These players exceed" in val_b or "WAITLISTED" in val_b.upper():
                continue

            # Count BYE slots towards group size but don't add to players list
            if val_b == "--BYE--":
                if not in_waitlist:
                    cur_group_slots += 1
                continue

            # Real player slot
            if not in_waitlist:
                cur_group_slots += 1

            is_seeded = (not in_waitlist) and bool(
                cell_b.font and cell_b.font.bold
            )
            entry = {"label": val_b, "is_seeded": is_seeded}
            players.append(entry)
            if is_seeded:
                seeded.append(entry)

        # Flush last group
        if cur_group_slots:
            group_sizes.append(cur_group_slots)

        if group_sizes:
            group_size = max(set(group_sizes), key=group_sizes.count)  # mode
            draw_size  = len(group_sizes) * group_size
        else:
            group_size = 0
            draw_size  = 0

        result[cat] = {
            "players":    players,
            "seeded":     seeded,
            "draw_size":  draw_size,
            "group_size": group_size,
        }

    return result


# ──────────────────────────────────────────────────────────────────
# SEEDING – BUILD BALANCED SEEDED GROUPS
# ──────────────────────────────────────────────────────────────────

def make_seeded_draw_groups(all_players, seeded_players, draw_size, group_size):
    """
    Place seeds at top/bottom of groups in a balanced pattern, then fill
    remaining slots with randomly shuffled non-seeded players / BYEs.

    Balanced seed placement order (N = num_groups):
      Tops    : G1, G2, …, GN          (seeds 1 … N)
      Bottoms : GN, G(N-1), …, G1      (seeds N+1 … 2N, reversed for balance)

    Args:
        all_players   – complete player list (dicts with 'label').
        seeded_players – seeded subset in seed-order (Seed 1 first).
        draw_size, group_size – same constraints as make_draw_groups.

    Returns:
        (groups, waitlisted)
    """
    if draw_size % group_size != 0:
        raise ValueError(
            f"group_size ({group_size}) must divide draw_size ({draw_size}) exactly."
        )

    num_groups = draw_size // group_size

    seeded_labels = {p["label"] for p in seeded_players}
    non_seeded    = [p for p in all_players if p["label"] not in seeded_labels]
    random.shuffle(non_seeded)

    # Waitlist excess non-seeded first (seeds always get a spot)
    total = len(seeded_players) + len(non_seeded)
    if total > draw_size:
        extra      = total - draw_size
        waitlisted = non_seeded[-extra:]
        non_seeded = non_seeded[:-extra]
    else:
        waitlisted = []

    # Seed slot positions: tops L→R, then bottoms R→L
    top_slots    = [(g, 0)               for g in range(num_groups)]
    bottom_slots = [(g, group_size - 1)  for g in range(num_groups - 1, -1, -1)]
    all_seed_slots = top_slots + bottom_slots

    # If more seeds than available slots, overflow seeds become regular players
    seeds_to_place  = seeded_players[:len(all_seed_slots)]
    overflow_seeds  = seeded_players[len(all_seed_slots):]
    non_seeded      = overflow_seeds + non_seeded
    random.shuffle(non_seeded)

    # Build groups as flat slot arrays (None = empty)
    groups_raw = [[None] * group_size for _ in range(num_groups)]

    # Place seeds in designated slots
    for seed, (g_idx, slot_idx) in zip(seeds_to_place, all_seed_slots):
        entry = dict(seed)
        entry["is_seeded"] = True
        groups_raw[g_idx][slot_idx] = entry

    # Fill remaining None slots with non-seeded players, then BYEs
    ns_iter   = iter(non_seeded)
    bye_entry = {"label": "--BYE--", "is_bye": True}

    for g in groups_raw:
        for i in range(len(g)):
            if g[i] is None:
                player = next(ns_iter, None)
                g[i]   = player if player is not None else bye_entry

    return groups_raw, waitlisted


# ──────────────────────────────────────────────────────────────────
# SEEDING – GENERATE SEEDED DRAW EXCEL
# ──────────────────────────────────────────────────────────────────

def generate_seeded_draw_excel(draw_configs):
    """
    Like generate_draw_excel but uses seeded placement.

    draw_configs: dict  cat → {
        "players":    list of player dicts (all, may include is_seeded flag),
        "seeded":     list of seeded player dicts in seed order (Seed 1 first),
        "draw_size":  int,
        "group_size": int,
    }
    Returns bytes.
    """
    wb = openpyxl.Workbook()
    sheet_cfg = [
        ("MS",    "TT 2026 Mens Singles"),
        ("MD",    "TT 2026 Mens Doubles"),
        ("Mixed", "TT 2026 Mixed Doubles"),
        ("WS",    "TT 2026 Womens Singles"),
        ("WD",    "TT 2026 Womens Doubles"),
    ]

    category_counts = []
    first = True

    for cat, title in sheet_cfg:
        cfg = draw_configs.get(cat)
        if not cfg or not cfg.get("players"):
            continue

        all_players = cfg["players"]
        seeded      = cfg.get("seeded", [])
        draw_size   = cfg["draw_size"]
        group_size  = cfg["group_size"]

        groups, waitlisted = make_seeded_draw_groups(
            all_players, seeded, draw_size, group_size
        )

        if first:
            ws    = wb.active
            ws.title = cat
            first = False
        else:
            ws = wb.create_sheet(title=cat)

        write_draw_sheet(ws, title, groups, waitlisted=waitlisted)

        byes = sum(1 for g in groups for p in g if isinstance(p, dict) and p.get("is_bye"))
        category_counts.append((
            title,
            len(all_players),
            draw_size,
            len(groups),
            byes,
            len(waitlisted),
        ))

    sum_ws = wb.create_sheet(title="Summary")
    write_summary_sheet(sum_ws, category_counts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# MAIN  (command-line entry point)
# ──────────────────────────────────────────────────────────────────
def main():
    print("=" * 64)
    print("  TT Cybage Internal 2026 – Draw Maker  (CLI mode)")
    print("=" * 64)

    # 1. Read nominations
    print(f"\n[1] Reading nominations from '{INPUT_FILE}' …")
    data, emp_lookup = read_nominations(INPUT_FILE)

    cat_labels = {
        "MS":    "Men's Singles",
        "MD":    "Men's Doubles",
        "Mixed": "Mixed Doubles",
        "WS":    "Women's Singles",
        "WD":    "Women's Doubles",
    }

    # 2. Auto-configure: use next power-of-2 draw size for each category
    print("\n[2] Draw Configuration  (auto: next power of 2)")
    draw_configs = {}
    for cat, label in cat_labels.items():
        players = data[cat]
        n = len(players)
        if n == 0:
            continue

        draw_size  = next_power_of_2(n)
        gs_options = valid_group_sizes(draw_size)
        group_size = 16 if 16 in gs_options else gs_options[-1]
        num_groups = draw_size // group_size
        byes       = draw_size - n

        print(f"    {label:22s}: {n:3d} entries  →  "
              f"draw {draw_size:4d}  "
              f"({num_groups} group(s) of {group_size}, {byes} BYE(s))")

        draw_configs[cat] = {
            "entries":    players,
            "draw_size":  draw_size,
            "group_size": group_size,
        }

    # 3. Generate
    print(f"\n[3] Generating '{OUTPUT_FILE}' …")
    excel_bytes = generate_draw_excel(draw_configs)

    with open(OUTPUT_FILE, "wb") as f:
        f.write(excel_bytes)

    print(f"[4] Done! Saved as: {OUTPUT_FILE}")
    print("\nTip: Run  streamlit run app.py  for the interactive UI.")
    print("=" * 64)


if __name__ == "__main__":
    main()
