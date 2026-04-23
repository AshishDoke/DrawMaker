"""
TTClash Draw Maker – Backend
Processes 'Selct Nomination TTClash.xls' (HTML table disguised as .xls)
and generates tournament draws for TT Clash format.

Categories:
    MS    = Men's Singles
    WS    = Women's Singles
    35MS  = 35+ Men's Singles
    MD    = Men's Doubles
    WD    = Women's Doubles
    Mixed = Mixed Doubles

Company-based draw:
    For MS, WS, 35MS categories the coordinator can opt for company-based
    grouping that tries to avoid placing same-company players in the same group.
    (The registration form collects Company/Organisation; this is used as the
    location/affiliation proxy since no City column exists in the form data.)
"""

import io
import random
from collections import defaultdict
from datetime import datetime
from html.parser import HTMLParser

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ────────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ────────────────────────────────────────────────────────────────────────────────

POWERS_OF_2 = [4, 8, 16, 32, 64, 128, 256]

TTCLASH_CATEGORIES = {
    "MS":    "Men's Singles",
    "WS":    "Women's Singles",
    "35MS":  "35+ Men's Singles",
    "MD":    "Men's Doubles",
    "WD":    "Women's Doubles",
    "Mixed": "Mixed Doubles",
}

# Categories that support company-based draw grouping
COMPANY_DRAW_CATEGORIES = {"MS", "WS", "35MS"}

PLAYERS_SHEET_NAME = "Players"


def next_power_of_2(n):
    """Smallest power of 2 >= n (minimum 4)."""
    p = 4
    while p < n:
        p *= 2
    return p


def valid_group_sizes(draw_size):
    """Return all valid group sizes for a given draw_size."""
    sizes = []
    p = 4
    while p <= draw_size:
        sizes.append(p)
        p *= 2
    return sizes


# ────────────────────────────────────────────────────────────────────────────────
# STYLES  (independent copy – same palette as draw_maker.py)
# ────────────────────────────────────────────────────────────────────────────────

TITLE_FONT       = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
TITLE_FILL       = PatternFill("solid", fgColor="1F4E79")
GROUP_FONT       = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
GROUP_FILL       = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL      = PatternFill("solid", fgColor="ED7D31")
PLAYER_FONT      = Font(name="Calibri", size=10)
PLAYER_FILL_A    = PatternFill("solid", fgColor="DEEAF1")
PLAYER_FILL_B    = PatternFill("solid", fgColor="FFFFFF")
BYE_FONT         = Font(name="Calibri", size=10, italic=True, color="808080")
SEEDED_FONT      = Font(name="Calibri", size=10, bold=True, color="1F4E79")
SEEDED_FILL_A    = PatternFill("solid", fgColor="D6E4F0")
SEEDED_FILL_B    = PatternFill("solid", fgColor="EBF3FB")
CENTER_ALIGN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LINK_FONT        = Font(name="Calibri", size=10, color="0563C1", underline="single")
SEEDED_LINK_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79", underline="single")


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ────────────────────────────────────────────────────────────────────────────────
# HTML PARSER
# ────────────────────────────────────────────────────────────────────────────────

class _HtmlTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self._row  = []
        self._cell = ""
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag in ("td", "th"):
            self._in_cell = True
            self._cell = ""

    def handle_endtag(self, tag):
        if tag in ("td", "th"):
            self._row.append(self._cell.strip())
            self._in_cell = False
        elif tag == "tr":
            if self._row:
                self.rows.append(self._row)
                self._row = []

    def handle_data(self, data):
        if self._in_cell:
            self._cell += data

    def handle_entityref(self, name):
        if self._in_cell and name == "nbsp":
            self._cell += " "


def _parse_html_table(source):
    """
    Parse a TTClash nomination file into a list of rows (each row = list of strings).
    Supports:
      - Real .xlsx files (detected by PK zip magic bytes)
      - HTML tables (.xls saved as HTML, plain .htm/.html)
    Raises ValueError if the file is a multi-frame HTML frameset wrapper.
    """
    import io as _io

    # Normalise input to raw bytes
    if isinstance(source, (bytes, bytearray)):
        raw = bytes(source)
    elif hasattr(source, "read"):
        raw = source.read()
        if isinstance(raw, str):
            raw = raw.encode("utf-8")
    else:
        with open(str(source), "rb") as f:
            raw = f.read()

    # ── Real .xlsx: PK zip magic ─────────────────────────────────────────────
    if raw[:4] == b"PK\x03\x04":
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for xlrow in ws.iter_rows(values_only=True):
            if any(v is not None for v in xlrow):
                rows.append([str(v).strip() if v is not None else "" for v in xlrow])
        wb.close()
        return rows

    # ── HTML table (.xls / .htm / .html) ────────────────────────────────────
    content = raw.decode("utf-8", errors="replace")
    parser = _HtmlTableParser()
    parser.feed(content)

    # Detect multi-frame Excel workbook (frameset wrapper with no table data)
    if not parser.rows:
        import re as _re
        m = _re.search(r'<link[^>]+id=["\']shLink["\'][^>]+href=["\']([^"\']+)["\']', content, _re.IGNORECASE)
        if m:
            companion = m.group(1).replace("%20", " ")
            raise ValueError(f"FRAMESET_WORKBOOK:{companion}")

    return parser.rows


# ────────────────────────────────────────────────────────────────────────────────
# STEP 1 – READ NOMINATIONS
# ────────────────────────────────────────────────────────────────────────────────

def _safe(row, idx):
    """Safely get a cell value from a row, returning '' on missing/None index."""
    if idx is None:
        return ""
    try:
        return (row[idx] or "").strip()
    except (IndexError, TypeError):
        return ""


def _make_doubles_pairs(raw_entries, seen_keys):
    """
    Deduplicate doubles pairs and build pair label dicts.

    raw_entries : list of dicts with keys:
        name, email, company, name_b, email_b, contact_b
    seen_keys   : set – modified in-place to track seen pair keys.

    Returns list of pair dicts with keys:
        label, email (primary player A), company (primary player A)
    """
    pairs = []
    for entry in raw_entries:
        email_a = entry["email"].lower().strip()
        email_b = entry["email_b"].lower().strip()

        if email_a and email_b:
            key = frozenset({email_a, email_b})
        else:
            key = (email_a, entry["name_b"].lower().strip())

        if key in seen_keys:
            continue
        seen_keys.add(key)

        name_a    = entry["name"]
        name_b    = entry["name_b"]
        company_a = entry.get("company", "")

        if name_b:
            contact_b = entry.get("contact_b", "")
            label_b   = f"{name_b} ({contact_b})" if contact_b else name_b
        else:
            label_b = "TBD"

        label_a    = f"{name_a} ({company_a})" if company_a else name_a
        pair_label = f"{label_a}  &  {label_b}"

        pairs.append({
            "label":   pair_label,
            "email":   email_a,   # primary email for hyperlink look-ups
            "company": company_a, # primary company for company-based grouping
        })

    return pairs


def read_ttclash_nominations(source):
    """
    Parse a TTClash .xls (HTML table) file and return categorised nominations.

    Returns:
        categories  : dict  cat → list of player dicts
            Singles (MS, WS, 35MS): label, name, email, company, mobile, gender, tshirt
            Doubles (MD, WD, Mixed): label, email (player A), company (player A)
        player_info : list of unique registrant dicts for the Players Info sheet
                      {regno, name, mobile, company, gender, tshirt, email}
    """
    rows = _parse_html_table(source)
    if not rows:
        return {cat: [] for cat in TTCLASH_CATEGORIES}, [], False

    header_row = rows[0]
    data_rows  = rows[1:]   # skip header row

    # Build a case-insensitive header → column-index map so the parser works
    # regardless of column order or whether extra columns (e.g. City) were added.
    col = {h.strip().lower(): i for i, h in enumerate(header_row)}

    # Column index helpers (falling back to legacy fixed positions when a
    # renamed / missing header leaves the key absent from col)
    c_regno   = col.get("regno",               0)
    c_name    = col.get("name",                1)
    c_company = col.get("company",             col.get("organisation", col.get("organization", 2)))
    c_email   = col.get("email",               3)
    c_mobile  = col.get("contactno",           col.get("mobile",       col.get("contact", 4)))
    c_gender  = col.get("gender",              5)
    c_tshirt  = col.get("tshirtsize",          col.get("tshirt",       9))
    c_city    = col.get("city")                # None when column absent

    c_ms      = col.get("menssingles",         col.get("ms",           10))
    c_ws      = col.get("womenssingles",       col.get("ws",           11))
    c_35ms    = col.get("35plusmenssingles",   col.get("35ms",         12))
    c_md      = col.get("mensdoubles",         col.get("md",           13))
    c_wd      = col.get("womensdoubles",       col.get("wd",           14))
    c_mixed   = col.get("mixeddoubles",        col.get("mixed",        15))

    c_med_name  = col.get("medname",           16)
    c_med_email = col.get("medemail",          17)
    c_med_cont  = col.get("medcontactno",      18)

    c_mid_name  = col.get("midname",           col.get("medname2",     22))
    c_mid_email = col.get("midemail",          col.get("medemail2",    23))
    c_mid_cont  = col.get("midcontactno",      col.get("medcontactno2",24))

    ms_list   = []
    ws_list   = []
    ms35_list = []
    md_raw    = []
    wd_raw    = []
    mix_raw   = []

    player_info_list = []
    seen_emails      = set()

    for row in data_rows:
        regno   = _safe(row, c_regno)
        name    = _safe(row, c_name)
        company = _safe(row, c_company)
        email   = _safe(row, c_email).lower()
        mobile  = _safe(row, c_mobile)
        gender  = _safe(row, c_gender)
        tshirt  = _safe(row, c_tshirt)
        city    = _safe(row, c_city)   # '' when c_city is None

        is_ms    = bool(_safe(row, c_ms))
        is_ws    = bool(_safe(row, c_ws))
        is_35ms  = bool(_safe(row, c_35ms))
        is_md    = bool(_safe(row, c_md))
        is_wd    = bool(_safe(row, c_wd))
        is_mixed = bool(_safe(row, c_mixed))

        med_name  = _safe(row, c_med_name)
        med_email = _safe(row, c_med_email).lower()
        med_cont  = _safe(row, c_med_cont)

        mid_name  = _safe(row, c_mid_name)
        mid_email = _safe(row, c_mid_email).lower()
        mid_cont  = _safe(row, c_mid_cont)

        if not name:
            continue

        # Collect unique registrant info (deduplicated by email)
        if email and email not in seen_emails:
            seen_emails.add(email)
            player_info_list.append({
                "regno":   regno,
                "name":    name,
                "mobile":  mobile,
                "company": company,
                "city":    city,
                "gender":  gender,
                "tshirt":  tshirt,
                "email":   email,
            })

        city_suffix  = f" [{city}]" if city else ""
        player_label = f"{name} ({company}){city_suffix}" if company else f"{name}{city_suffix}"

        singles_base = {
            "label":   player_label,
            "name":    name,
            "email":   email,
            "company": company,
            "mobile":  mobile,
            "gender":  gender,
            "tshirt":  tshirt,
            "city":    city,
        }

        if is_ms:
            ms_list.append(singles_base.copy())
        if is_ws:
            ws_list.append(singles_base.copy())
        if is_35ms:
            ms35_list.append(singles_base.copy())

        if is_md:
            md_raw.append({
                "name": name, "email": email, "company": company,
                "name_b": med_name, "email_b": med_email, "contact_b": med_cont,
            })

        if is_wd:
            # WD uses the same MEDName/MEDEmail partner fields
            wd_raw.append({
                "name": name, "email": email, "company": company,
                "name_b": med_name, "email_b": med_email, "contact_b": med_cont,
            })

        if is_mixed:
            mix_raw.append({
                "name": name, "email": email, "company": company,
                "name_b": mid_name, "email_b": mid_email, "contact_b": mid_cont,
            })

    md_seen  = set()
    wd_seen  = set()
    mix_seen = set()

    categories = {
        "MS":    ms_list,
        "WS":    ws_list,
        "35MS":  ms35_list,
        "MD":    _make_doubles_pairs(md_raw,  md_seen),
        "WD":    _make_doubles_pairs(wd_raw,  wd_seen),
        "Mixed": _make_doubles_pairs(mix_raw, mix_seen),
    }

    has_city_data = (
        c_city is not None
        and any(p.get("city") for p in ms_list + ws_list + ms35_list)
    )

    return categories, player_info_list, has_city_data


# ────────────────────────────────────────────────────────────────────────────────
# DRAW GROUP BUILDERS
# ────────────────────────────────────────────────────────────────────────────────

def _interleave_by_field(players, field):
    """
    Reorder players so those sharing the same value for `field` are spread as
    evenly as possible across the resulting list.
    Uses a greedy largest-bucket algorithm:
      – group players by the given field value
      – on each round, take one player from each non-empty bucket (sorted by
        descending remaining size so the largest groups are spread first)
    """
    buckets = defaultdict(list)
    for p in players:
        key = (p.get(field) or "").strip().lower() or "_no_data"
        buckets[key].append(p)

    bucket_list = list(buckets.values())
    for b in bucket_list:
        random.shuffle(b)

    result = []
    while bucket_list:
        bucket_list.sort(key=len, reverse=True)
        next_round = []
        for b in bucket_list:
            if b:
                result.append(b.pop(0))
            if b:
                next_round.append(b)
        bucket_list = next_round

    return result


def make_ttclash_draw_groups(players, draw_size, group_size, by_city=False, by_company=False):
    """
    Build draw groups from a player list.

    by_city=True     – interleave by City (if city data present in player dicts).
    by_company=True  – interleave by Company.
    When both are True, City takes priority; falls back to Company if no city data.

    Returns (groups, waitlisted).
    """
    if draw_size % group_size != 0:
        raise ValueError(
            f"group_size ({group_size}) must divide draw_size ({draw_size}) exactly."
        )

    if by_city and any(p.get("city") for p in players):
        ordered = _interleave_by_field(players, "city")
    elif by_company:
        ordered = _interleave_by_field(players, "company")
    else:
        ordered = list(players)
        random.shuffle(ordered)

    n = len(ordered)
    if n > draw_size:
        waitlisted = ordered[draw_size:]
        active     = ordered[:draw_size]
    else:
        waitlisted = []
        active     = ordered[:]

    bye_entry = {"label": "--BYE--", "is_bye": True}
    all_slots = active + [bye_entry] * max(0, draw_size - len(active))
    random.shuffle(all_slots)  # distribute BYEs randomly across all groups

    num_groups = draw_size // group_size
    groups = [
        all_slots[i * group_size: (i + 1) * group_size]
        for i in range(num_groups)
    ]
    return groups, waitlisted


def make_seeded_draw_groups(all_players, seeded_players, draw_size, group_size):
    """
    Balanced seeded placement – identical algorithm to draw_maker.py.
    Seeds placed at top/bottom of groups in a balanced interleaved pattern.
    """
    if draw_size % group_size != 0:
        raise ValueError(
            f"group_size ({group_size}) must divide draw_size ({draw_size}) exactly."
        )

    num_groups    = draw_size // group_size
    seeded_labels = {p["label"] for p in seeded_players}
    non_seeded    = [p for p in all_players if p["label"] not in seeded_labels]
    random.shuffle(non_seeded)

    total = len(seeded_players) + len(non_seeded)
    if total > draw_size:
        extra      = total - draw_size
        waitlisted = non_seeded[-extra:]
        non_seeded = non_seeded[:-extra]
    else:
        waitlisted = []

    top_slots    = [(g, 0)              for g in range(num_groups)]
    bottom_slots = [(g, group_size - 1) for g in range(num_groups - 1, -1, -1)]
    all_seed_slots = top_slots + bottom_slots

    seeds_to_place = seeded_players[:len(all_seed_slots)]
    overflow_seeds = seeded_players[len(all_seed_slots):]
    non_seeded     = overflow_seeds + non_seeded
    random.shuffle(non_seeded)

    groups_raw = [[None] * group_size for _ in range(num_groups)]
    for seed, (g_idx, slot_idx) in zip(seeds_to_place, all_seed_slots):
        entry = dict(seed)
        entry["is_seeded"] = True
        groups_raw[g_idx][slot_idx] = entry

    ns_iter   = iter(non_seeded)
    bye_entry = {"label": "--BYE--", "is_bye": True}

    # Count empty (non-seed) slots and pad the non-seeded pool with BYEs so
    # they land in random positions rather than always at the end.
    num_open_slots = sum(1 for g in groups_raw for slot in g if slot is None)
    byes_needed    = max(0, num_open_slots - len(non_seeded))
    if byes_needed:
        padded = non_seeded + [bye_entry] * byes_needed
        random.shuffle(padded)
        ns_iter = iter(padded)

    for g in groups_raw:
        for i in range(len(g)):
            if g[i] is None:
                g[i] = next(ns_iter, bye_entry)

    return groups_raw, waitlisted


# ────────────────────────────────────────────────────────────────────────────────
# PLAYERS INFO SHEET
# ────────────────────────────────────────────────────────────────────────────────

def write_player_info_sheet(ws, player_info):
    """
    Write the Players info sheet.
    Returns {email.lower(): row_number} for hyperlink building.
    """
    col_widths  = [8, 30, 16, 30, 10, 16, 32]
    col_headers = ["Reg#", "Name", "Mobile", "Company / Org", "Gender", "T-Shirt Size", "Email"]

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    title = ws.cell(1, 1, "TT Clash 2026 \u2013 Player Information")
    title.font      = TITLE_FONT
    title.fill      = TITLE_FILL
    title.alignment = CENTER_ALIGN
    ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")

    ws.row_dimensions[2].height = 18
    for col, hdr in enumerate(col_headers, start=1):
        c = ws.cell(2, col, hdr)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN

    email_row_map = {}
    for r, info in enumerate(player_info, start=3):
        ws.row_dimensions[r].height = 16
        fill = PLAYER_FILL_A if r % 2 == 0 else PLAYER_FILL_B
        vals = [
            info.get("regno",   ""),
            info.get("name",    ""),
            info.get("mobile",  ""),
            info.get("company", ""),
            info.get("gender",  ""),
            info.get("tshirt",  ""),
            info.get("email",   ""),
        ]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(r, col, val)
            c.font      = PLAYER_FONT
            c.fill      = fill
            c.alignment = CENTER_ALIGN if col == 1 else LEFT_ALIGN
            c.border    = _thin_border()

        email = (info.get("email") or "").lower().strip()
        if email:
            email_row_map[email] = r

    ws.freeze_panes = "A3"
    return email_row_map


# ────────────────────────────────────────────────────────────────────────────────
# DRAW SHEET WRITER
# ────────────────────────────────────────────────────────────────────────────────

def write_ttclash_draw_sheet(ws, title, groups, waitlisted=None,
                              player_row_map=None, info_sheet_name="Players"):
    """
    Write a TTClash category draw sheet.

    player_row_map    : dict { email.lower() \u2192 row_number_in_Players_sheet }
                        If provided, a clickable \"...\" is placed in column A for each
                        non-BYE player that navigates to their row in the Players sheet.
                        The player name in column B is always plain text.
    info_sheet_name   : name of the Players sheet (default "Players").
    """
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 5
    ws.column_dimensions["J"].width = 33
    ws.column_dimensions["K"].width = 5
    ws.column_dimensions["L"].width = 5
    ws.column_dimensions["M"].width = 36

    # Row 1 – title + stage headers
    ws.row_dimensions[1].height = 24
    title_cell = ws.cell(1, 1, title)
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.merge_cells("A1:G1")

    for col, label in [(9, "QF"), (11, "SF"), (13, "Finals")]:
        c = ws.cell(1, col, label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN

    current_row   = 3
    player_number = 1
    group_mid_rows = []

    for g_idx, group in enumerate(groups):
        n_in_group = len(group)

        # Group header
        ws.row_dimensions[current_row].height = 18
        hdr = ws.cell(current_row, 1, f"------ GROUP - {g_idx + 1} -----")
        hdr.font      = GROUP_FONT
        hdr.fill      = GROUP_FILL
        hdr.alignment = LEFT_ALIGN
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=7,
        )
        current_row += 1
        current_row += 1   # blank row after header

        first_player_row = current_row

        for p_idx, player in enumerate(group):
            ws.row_dimensions[current_row].height = 17

            label     = player.get("label", str(player)) if isinstance(player, dict) else str(player)
            is_bye    = player.get("is_bye",    False) if isinstance(player, dict) else False
            is_seeded = player.get("is_seeded", False) if isinstance(player, dict) else False
            email     = (player.get("email", "") or "").lower() if isinstance(player, dict) else ""

            name_cell = ws.cell(current_row, 2, label)
            num_cell  = ws.cell(current_row, 1)

            if is_bye:
                fill      = PatternFill("solid", fgColor="F2F2F2")
                name_font = BYE_FONT
                num_cell.value = ""
                num_cell.font  = BYE_FONT
            else:
                if is_seeded:
                    fill      = SEEDED_FILL_A if p_idx % 2 == 0 else SEEDED_FILL_B
                    name_font = SEEDED_FONT
                    dots_font = Font(name="Calibri", size=9, bold=True, color="1F4E79")
                else:
                    fill      = PLAYER_FILL_A if p_idx % 2 == 0 else PLAYER_FILL_B
                    name_font = PLAYER_FONT
                    dots_font = Font(name="Calibri", size=9, color="0563C1")

                # Col A: HYPERLINK formula "..." → Players sheet, or plain number
                if player_row_map and email:
                    prow = player_row_map.get(email)
                    if prow:
                        sref = (
                            f"'{info_sheet_name}'!A{prow}"
                            if " " in info_sheet_name
                            else f"{info_sheet_name}!A{prow}"
                        )
                        num_cell.value = f'=HYPERLINK("#{sref}","...")'
                        num_cell.font  = dots_font
                    else:
                        num_cell.value = player_number
                        num_cell.font  = Font(name="Calibri", size=10, bold=True)
                else:
                    num_cell.value = player_number
                    num_cell.font  = Font(name="Calibri", size=10, bold=True)

            for c in [num_cell, name_cell]:
                c.fill      = fill
                c.border    = _thin_border()
                c.alignment = CENTER_ALIGN if c.column == 1 else LEFT_ALIGN

            name_cell.font = name_font   # always plain text, no hyperlink styling

            if not is_bye:
                player_number += 1
            current_row += 1

        last_player_row = current_row - 1
        group_mid       = (first_player_row + last_player_row) // 2

        # QF bracket placeholder cells
        if n_in_group >= 4:
            half    = n_in_group // 2
            qf1_row = (first_player_row + first_player_row + half - 1) // 2
            qf2_row = (first_player_row + half + last_player_row) // 2
        else:
            qf1_row = group_mid
            qf2_row = group_mid

        for qf_row in set([qf1_row, qf2_row]):
            c = ws.cell(qf_row, 8)
            c.border    = Border(bottom=Side(style="medium", color="2E75B6"))
            c.alignment = LEFT_ALIGN
            c.font      = Font(name="Calibri", size=10, italic=True, color="808080")

        sf_cell = ws.cell(group_mid, 10)
        sf_cell.border    = Border(bottom=Side(style="medium", color="ED7D31"))
        sf_cell.alignment = LEFT_ALIGN
        sf_cell.font      = Font(name="Calibri", size=10, italic=True, color="808080")

        group_mid_rows.append(group_mid)
        current_row += 1   # blank separator

    # Finals placeholder
    if group_mid_rows:
        finals_row = (group_mid_rows[0] + group_mid_rows[-1]) // 2
        f_cell = ws.cell(finals_row, 13)
        f_cell.border    = Border(bottom=Side(style="thick", color="1F4E79"))
        f_cell.alignment = LEFT_ALIGN
        f_cell.font      = Font(name="Calibri", size=10, bold=True, italic=True, color="1F4E79")

    # Waitlisted section
    if waitlisted:
        current_row += 2

        ws.row_dimensions[current_row].height = 20
        wl_hdr = ws.cell(current_row, 1, "\u2500\u2500 WAITLISTED / RESERVE PLAYERS \u2500\u2500")
        wl_hdr.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        wl_hdr.fill      = PatternFill("solid", fgColor="595959")
        wl_hdr.alignment = LEFT_ALIGN
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=7,
        )
        current_row += 1

        ws.row_dimensions[current_row].height = 14
        note = ws.cell(current_row, 2,
                       "These players exceed the draw size. "
                       "Tournament coordinator to decide eligibility.")
        note.font      = Font(name="Calibri", size=9, italic=True, color="595959")
        note.alignment = LEFT_ALIGN
        current_row += 1

        for wl_player in waitlisted:
            ws.row_dimensions[current_row].height = 17
            lbl = (wl_player.get("label", str(wl_player))
                   if isinstance(wl_player, dict) else str(wl_player))
            c = ws.cell(current_row, 2, lbl)
            c.font      = Font(name="Calibri", size=10, italic=True, color="595959")
            c.fill      = PatternFill("solid", fgColor="FFF2CC")
            c.border    = _thin_border()
            c.alignment = LEFT_ALIGN
            current_row += 1

    ws.freeze_panes = "A2"
    return current_row


# ────────────────────────────────────────────────────────────────────────────────
# SUMMARY SHEET
# ────────────────────────────────────────────────────────────────────────────────

def write_ttclash_summary_sheet(ws, category_counts):
    """
    category_counts: list of (title, entries, draw_size, num_groups, byes, waitlisted_count)
    """
    col_widths  = [36, 12, 12, 10, 8, 14]
    col_headers = ["Category", "Entries", "Draw Size", "Groups", "BYEs", "Waitlisted"]

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    title = ws.cell(1, 1, "TT Clash 2026 \u2013 Draw Summary")
    title.font      = TITLE_FONT
    title.fill      = TITLE_FILL
    title.alignment = CENTER_ALIGN
    ws.merge_cells(f"A1:{get_column_letter(len(col_headers))}1")

    ws.row_dimensions[2].height = 18
    for col, hdr in enumerate(col_headers, start=1):
        c = ws.cell(2, col, hdr)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER_ALIGN

    for r, row_data in enumerate(category_counts, start=3):
        title_val, entries, draw_size, n_groups, byes, waitlisted_cnt = row_data
        ws.row_dimensions[r].height = 17
        fill = PLAYER_FILL_A if r % 2 == 0 else PLAYER_FILL_B
        for col, val in enumerate(
            [title_val, entries, draw_size, n_groups, byes, waitlisted_cnt], start=1
        ):
            c = ws.cell(r, col, val)
            c.font      = PLAYER_FONT
            c.fill      = fill
            c.alignment = CENTER_ALIGN
            c.border    = _thin_border()

    note_row = len(category_counts) + 5
    ws.row_dimensions[note_row].height = 15
    note = ws.cell(note_row, 1,
                   f"Draw generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}")
    note.font = Font(name="Calibri", size=9, italic=True, color="808080")


# ────────────────────────────────────────────────────────────────────────────────
# GENERATE DRAW EXCEL
# ────────────────────────────────────────────────────────────────────────────────

_SHEET_CFG = [
    ("MS",    "TT Clash 2026 \u2013 Men's Singles"),
    ("WS",    "TT Clash 2026 \u2013 Women's Singles"),
    ("35MS",  "TT Clash 2026 \u2013 35+ Men's Singles"),
    ("MD",    "TT Clash 2026 \u2013 Men's Doubles"),
    ("WD",    "TT Clash 2026 \u2013 Women's Doubles"),
    ("Mixed", "TT Clash 2026 \u2013 Mixed Doubles"),
]


def generate_ttclash_draw_excel(draw_configs, player_info=None):
    """
    Build the TTClash draw Excel workbook and return its bytes.

    draw_configs : dict  cat \u2192 {
        "entries"        : list of player dicts (each with 'label' and 'email'),
        "draw_size"      : int (power of 2),
        "group_size"     : int (even, power of 2, divides draw_size),
        "draw_by_company": bool – if True, interleave by Company (MS/WS/35MS only),
    }
    player_info  : list of registrant info dicts (for the Players sheet).
                   If provided and non-empty, a "Players" sheet is added and
                   player name cells in draw sheets will link to it.

    Returns bytes of the generated Excel file.
    """
    wb    = openpyxl.Workbook()
    first = True

    # Build Players info sheet first so we have the email\u2192row map
    player_row_map = {}
    if player_info:
        pws = wb.active
        pws.title = PLAYERS_SHEET_NAME
        first = False
        player_row_map = write_player_info_sheet(pws, player_info)

    category_counts = []

    for cat, title in _SHEET_CFG:
        if cat not in draw_configs or not draw_configs[cat].get("entries"):
            continue

        cfg        = draw_configs[cat]
        entries    = cfg["entries"]
        draw_size  = cfg["draw_size"]
        group_size = cfg["group_size"]
        by_city    = cfg.get("draw_by_city",    False) and cat in COMPANY_DRAW_CATEGORIES
        by_company = cfg.get("draw_by_company", False) and cat in COMPANY_DRAW_CATEGORIES

        players = [dict(p) for p in entries]
        groups, waitlisted = make_ttclash_draw_groups(
            players, draw_size, group_size, by_city=by_city, by_company=by_company
        )

        if first:
            ws = wb.active
            ws.title = cat
            first = False
        else:
            ws = wb.create_sheet(title=cat)

        write_ttclash_draw_sheet(
            ws, title, groups, waitlisted=waitlisted,
            player_row_map=player_row_map if player_info else None,
            info_sheet_name=PLAYERS_SHEET_NAME,
        )

        byes = draw_size - min(len(players), draw_size)
        category_counts.append((
            title,
            len(entries),
            draw_size,
            len(groups),
            byes,
            len(waitlisted),
        ))

    sum_ws = wb.create_sheet(title="Summary")
    write_ttclash_summary_sheet(sum_ws, category_counts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────────────────────────
# READ EXISTING DRAW FILE  (for re-seeding)
# ────────────────────────────────────────────────────────────────────────────────

_TTCLASH_DRAW_SHEETS = ["MS", "WS", "35MS", "MD", "WD", "Mixed"]


def read_ttclash_draw_file(filepath_or_bytes):
    """
    Parse a TTClash draw Excel (generated by this module).
    Detects player names in bold as seeded players.
    Also auto-detects draw_size and group_size from the sheet structure.

    Returns:
        dict  cat \u2192 {"players": [...], "seeded": [...],
                          "draw_size": int, "group_size": int}
    """
    wb = openpyxl.load_workbook(filepath_or_bytes)
    result = {}

    for cat in _TTCLASH_DRAW_SHEETS:
        if cat not in wb.sheetnames:
            result[cat] = {"players": [], "seeded": [], "draw_size": 0, "group_size": 0}
            continue

        ws          = wb[cat]
        players     = []
        seeded      = []
        in_waitlist = False
        group_sizes = []
        cur_slots   = 0

        for row in range(2, ws.max_row + 1):
            cell_a = ws.cell(row, 1)
            cell_b = ws.cell(row, 2)
            val_a  = str(cell_a.value or "").strip()
            val_b  = str(cell_b.value or "").strip()

            if "WAITLISTED" in val_a.upper():
                if cur_slots:
                    group_sizes.append(cur_slots)
                    cur_slots = 0
                in_waitlist = True
                continue

            if val_a.startswith("------"):
                if cur_slots:
                    group_sizes.append(cur_slots)
                    cur_slots = 0
                continue

            if not val_b:
                continue
            if "These players exceed" in val_b or "WAITLISTED" in val_b.upper():
                continue

            if val_b == "--BYE--":
                if not in_waitlist:
                    cur_slots += 1
                continue

            if not in_waitlist:
                cur_slots += 1

            is_seeded = (not in_waitlist) and bool(
                cell_b.font and cell_b.font.bold
            )
            entry = {"label": val_b, "is_seeded": is_seeded}
            players.append(entry)
            if is_seeded:
                seeded.append(entry)

        if cur_slots:
            group_sizes.append(cur_slots)

        if group_sizes:
            group_size = max(set(group_sizes), key=group_sizes.count)
            draw_size  = len(group_sizes) * group_size
        else:
            group_size = 0
            draw_size  = 0

        result[cat] = {
            "players":    players,
            "seeded":     seeded,
            "draw_size":  draw_size,
            "group_size": group_size,
        }

    return result


# ────────────────────────────────────────────────────────────────────────────────
# GENERATE SEEDED DRAW EXCEL
# ────────────────────────────────────────────────────────────────────────────────

def generate_ttclash_seeded_draw_excel(draw_configs):
    """
    Build a seeded TTClash draw Excel workbook and return its bytes.

    draw_configs : dict  cat \u2192 {
        "players":    list of all player dicts (each with 'label'),
        "seeded":     list of seeded player dicts in seed order (Seed 1 first),
        "draw_size":  int,
        "group_size": int,
    }
    Returns bytes.
    """
    wb    = openpyxl.Workbook()
    first = True
    category_counts = []

    for cat, title in _SHEET_CFG:
        cfg = draw_configs.get(cat)
        if not cfg or not cfg.get("players"):
            continue

        all_players = cfg["players"]
        seeded      = cfg.get("seeded", [])
        draw_size   = cfg["draw_size"]
        group_size  = cfg["group_size"]

        groups, waitlisted = make_seeded_draw_groups(
            all_players, seeded, draw_size, group_size
        )

        if first:
            ws = wb.active
            ws.title = cat
            first = False
        else:
            ws = wb.create_sheet(title=cat)

        write_ttclash_draw_sheet(ws, title, groups, waitlisted=waitlisted)

        byes = sum(
            1 for g in groups
            for p in g if isinstance(p, dict) and p.get("is_bye")
        )
        category_counts.append((
            title,
            len(all_players),
            draw_size,
            len(groups),
            byes,
            len(waitlisted),
        ))

    sum_ws = wb.create_sheet(title="Summary")
    write_ttclash_summary_sheet(sum_ws, category_counts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
